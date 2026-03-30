from uuid import UUID
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.models.import_row import ImportRow
from app.models.assignment import UserCourseAssignment
from app.models.attempt import AttemptStatus
from app.models.user import User


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")


def _style_header(ws, headers: list[str]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


class ExportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_logins_passwords(self, batch_id: UUID) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Доступы"
        _style_header(ws, ["№", "ФИО", "Логин", "Пароль", "Организация", "Должность", "Назначенные курсы"])

        result = await self.db.execute(
            select(ImportRow)
            .options(selectinload(ImportRow.user))
            .where(ImportRow.batch_id == batch_id, ImportRow.user_id.isnot(None))
            .order_by(ImportRow.row_number)
        )
        rows = result.scalars().all()

        for i, row in enumerate(rows, 1):
            u = row.user
            nd = row.normalized_data or {}
            assigned = nd.get("assigned", [])
            ws.append([
                i,
                u.full_name if u else nd.get("full_name", ""),
                u.login if u else "",
                nd.get("password", ""),
                nd.get("organization", ""),
                u.position_raw if u else nd.get("position", ""),
                ", ".join(assigned) if isinstance(assigned, list) else str(assigned),
            ])

        return wb

    async def export_all_results(self) -> Workbook:
        return await self._build_results_workbook(batch_id=None)

    async def export_batch_results(self, batch_id: UUID) -> Workbook:
        return await self._build_results_workbook(batch_id=batch_id)

    async def _build_results_workbook(self, batch_id: UUID | None) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты"
        _style_header(ws, [
            "№", "ФИО", "Логин", "Организация", "Должность",
            "Дисциплина", "Курс", "Статус",
            "Лучший результат (%)", "Сдал?", "Дата завершения"
        ])

        STATUS_RU = {
            "assigned": "Назначен",
            "in_progress": "В процессе",
            "passed": "Сдан",
            "failed": "Не сдан",
        }

        q = (
            select(UserCourseAssignment)
            .options(
                selectinload(UserCourseAssignment.user).selectinload(User.organization),
                selectinload(UserCourseAssignment.course),
                selectinload(UserCourseAssignment.discipline),
                selectinload(UserCourseAssignment.attempts),
            )
            .order_by(UserCourseAssignment.assigned_at)
        )
        if batch_id:
            q = q.where(UserCourseAssignment.batch_id == batch_id)

        result = await self.db.execute(q)
        assignments = result.scalars().all()

        for i, a in enumerate(assignments, 1):
            u = a.user
            best_attempt = None
            if a.attempts:
                completed = [x for x in a.attempts if x.status == AttemptStatus.completed]
                if completed:
                    best_attempt = max(completed, key=lambda x: x.score_percent or 0)

            status_str = a.status.value if hasattr(a.status, "value") else str(a.status)
            ws.append([
                i,
                u.full_name if u else "",
                u.login if u else "",
                u.organization.name if (u and u.organization) else "",
                u.position_raw if u else "",
                a.discipline.name if a.discipline else "",
                a.course.name if a.course else "",
                STATUS_RU.get(status_str, status_str),
                best_attempt.score_percent if best_attempt else "",
                "Да" if (best_attempt and best_attempt.passed) else ("Нет" if best_attempt else ""),
                a.completed_at.strftime("%d.%m.%Y") if a.completed_at else "",
            ])

        return wb
