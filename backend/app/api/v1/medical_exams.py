"""
Medical exam API.

Admin routes (mounted at /admin/medical-exams):
  POST /import          — upload Excel from clinic, parse and save records
  GET  /                — list all records (filterable by org_id)
  DELETE /{id}          — delete single record

Learner routes (mounted at /learner/medical-exams):
  GET  /                — view own medical exam records
"""
import os
import uuid
import io
from typing import Optional
from datetime import datetime, date

from pydantic import BaseModel
from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Query
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.api.deps import CurrentAdmin, CurrentSuperAdmin, CurrentLearner, DB
from app.core.config import settings
from app.models.medical_exam import MedicalExam
from app.models.organization import Organization
from app.models.user import User

router = APIRouter()          # /admin/medical-exams
learner_router = APIRouter()  # /learner/medical-exams


# ── Schemas ───────────────────────────────────────────────────────────────────

class MedExamOut(BaseModel):
    id: uuid.UUID
    user_id: Optional[uuid.UUID]
    full_name: str
    birth_date: Optional[date]
    gender: Optional[str]
    workplace: Optional[str]
    position: Optional[str]
    icd10_group: Optional[str]
    fit_for_work: Optional[bool]
    exam_date: Optional[date]
    source_file: Optional[str]
    imported_at: datetime

    class Config:
        from_attributes = True


class AssignUserIn(BaseModel):
    user_id: uuid.UUID


# ── Excel parser ──────────────────────────────────────────────────────────────

def _parse_med_excel(content: bytes, filename: str) -> list[dict]:
    """
    Parse clinic's medical examination Excel report.

    Expected columns (by header text, case-insensitive):
      ФИО                                     → full_name
      Дата рождения                            → birth_date
      Пол                                     → gender
      Объект (или участок)                    → workplace
      Занимаемая должность                    → position
      Класс заболевания / МКБ                 → icd10_group
      Профпри / Годен                         → fit_for_work  ('+' = True)
      Дата прохождения                        → exam_date

    Falls back to column-position heuristic if headers aren't found.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Excel файл пуст")

    # ── Find header row ───────────────────────────────────────────────────────
    header_row_idx = None
    col_map: dict[str, int] = {}

    FIELD_KEYWORDS = {
        "full_name":    ["фио", "ф.и.о", "фамилия"],
        "birth_date":   ["дата рождения", "дата_рожд", "год рождения"],
        "gender":       ["пол"],
        "workplace":    ["объект", "участок"],
        "position":     ["должность", "занимаемая"],
        "icd10_group":  ["мкб", "класс заболев", "диспансер"],
        "fit_for_work": ["годен", "профпри"],
        "exam_date":    ["дата прохождения", "дата осмотра", "дата медосмотра"],
    }

    for r_idx, row in enumerate(rows):
        row_texts = [str(c).lower().strip() if c else "" for c in row]
        # A header row should contain "фио" or "фамилия"
        if any("фио" in t or "фамилия" in t for t in row_texts):
            header_row_idx = r_idx
            # Map field → column index
            for field, keywords in FIELD_KEYWORDS.items():
                for c_idx, cell_text in enumerate(row_texts):
                    if any(kw in cell_text for kw in keywords):
                        if field not in col_map:
                            col_map[field] = c_idx
            break

    # ── Fallback: use fixed column positions (B=1,C=2,D=3,E=4,F=5,I=8,K=10,U=20)
    # Based on the known UOC-2 template structure
    if header_row_idx is None or "full_name" not in col_map:
        header_row_idx = -1  # no header, data starts at row 0 (skip row 0 = title)
        # Find first row with a non-empty second cell that looks like a name
        for r_idx, row in enumerate(rows):
            if r_idx == 0:
                continue  # skip title
            cell_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if cell_b and cell_b not in ("ФИО", "Ф.И.О.", "№") and len(cell_b.split()) >= 2:
                header_row_idx = r_idx - 1  # one before data
                break
        col_map = {
            "full_name":    1,   # B
            "birth_date":   2,   # C
            "gender":       3,   # D
            "workplace":    4,   # E
            "position":     5,   # F
            "icd10_group":  8,   # I
            "fit_for_work": 10,  # K
            "exam_date":    20,  # U
        }

    # ── Parse data rows ───────────────────────────────────────────────────────
    records = []
    for row in rows[header_row_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue  # skip empty rows

        def get(field: str):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        full_name = get("full_name")
        if not full_name or len(full_name) < 2:
            continue  # skip rows without a name

        # Parse date fields
        def parse_date(val) -> Optional[date]:
            if val is None:
                return None
            if isinstance(val, (date, datetime)):
                return val.date() if isinstance(val, datetime) else val
            s = str(val).strip()
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    pass
            return None

        bd_raw = row[col_map["birth_date"]] if "birth_date" in col_map and col_map["birth_date"] < len(row) else None
        ed_raw = row[col_map["exam_date"]] if "exam_date" in col_map and col_map["exam_date"] < len(row) else None

        # fit_for_work: "+" or "Годен" → True; empty / "-" → False; None → None
        fw_raw = get("fit_for_work")
        if fw_raw is None or fw_raw == "":
            fit = None
        else:
            fit = fw_raw.strip() in ("+", "годен", "годна", "да", "yes", "1", "true")

        records.append({
            "full_name":    full_name,
            "birth_date":   parse_date(bd_raw),
            "gender":       get("gender"),
            "workplace":    get("workplace"),
            "position":     get("position"),
            "icd10_group":  get("icd10_group"),
            "fit_for_work": fit,
            "exam_date":    parse_date(ed_raw),
            "source_file":  filename,
        })

    return records


# ── Admin: Import ─────────────────────────────────────────────────────────────

@router.post("/import", status_code=201)
async def import_med_exams(
    db: DB,
    admin: CurrentSuperAdmin,
    organization_id: Optional[uuid.UUID] = Form(None),
    file: UploadFile = File(...),
):
    """Upload Excel file from clinic and import medical exam records."""
    content = await file.read()
    if not content:
        raise HTTPException(400, "Файл пустой")

    records = _parse_med_excel(content, file.filename or "import.xlsx")
    if not records:
        raise HTTPException(400, "В файле не найдено строк с данными. Проверьте формат Excel.")

    # Try to match full_name → user_id
    all_users = (await db.execute(select(User))).scalars().all()
    name_to_user: dict[str, User] = {u.full_name.lower().strip(): u for u in all_users}

    saved = 0
    unmatched_pairs: list[tuple[MedicalExam, dict]] = []
    for rec in records:
        matched_user = name_to_user.get(rec["full_name"].lower().strip())
        exam = MedicalExam(
            user_id=matched_user.id if matched_user else None,
            organization_id=organization_id,
            **{k: v for k, v in rec.items()},
        )
        db.add(exam)
        if not matched_user:
            unmatched_pairs.append((exam, rec))
        saved += 1

    await db.commit()

    # id is Python-generated (uuid.uuid4) so it's available immediately
    unmatched_out = [
        {
            "id": str(exam.id),
            "full_name": rec["full_name"],
            "exam_date": str(rec["exam_date"]) if rec["exam_date"] else None,
            "position": rec.get("position"),
        }
        for exam, rec in unmatched_pairs
    ]

    return {"imported": saved, "total_parsed": len(records), "unmatched": unmatched_out}


# ── Admin: Assign user ────────────────────────────────────────────────────────

@router.patch("/{exam_id}/assign-user")
async def assign_user_to_exam(
    exam_id: uuid.UUID,
    body: AssignUserIn,
    db: DB,
    admin: CurrentSuperAdmin,
):
    """Manually link a medical exam record to a system user."""
    result = await db.execute(select(MedicalExam).where(MedicalExam.id == exam_id))
    exam = result.scalar_one_or_none()
    if not exam:
        raise HTTPException(404, "Запись медосмотра не найдена")
    user_result = await db.execute(select(User).where(User.id == body.user_id))
    if not user_result.scalar_one_or_none():
        raise HTTPException(404, "Пользователь не найден")
    exam.user_id = body.user_id
    await db.commit()
    return {"ok": True}


# ── Admin: List ───────────────────────────────────────────────────────────────

@router.get("", response_model=list[MedExamOut])
async def list_med_exams(
    db: DB,
    admin: CurrentAdmin,
    organization_id: Optional[uuid.UUID] = Query(None),
):
    q = select(MedicalExam).order_by(MedicalExam.exam_date.desc(), MedicalExam.full_name)
    if organization_id:
        q = q.where(MedicalExam.organization_id == organization_id)
    result = await db.execute(q)
    return result.scalars().all()


# ── Admin: Delete ─────────────────────────────────────────────────────────────

@router.delete("/{exam_id}", status_code=204)
async def delete_med_exam(exam_id: uuid.UUID, db: DB, admin: CurrentSuperAdmin):
    result = await db.execute(select(MedicalExam).where(MedicalExam.id == exam_id))
    exam = result.scalar_one_or_none()
    if not exam:
        raise HTTPException(404, "Запись не найдена")
    await db.delete(exam)
    await db.commit()


# ── Learner: own records ──────────────────────────────────────────────────────

@learner_router.get("", response_model=list[MedExamOut])
async def my_med_exams(learner: CurrentLearner, db: DB):
    result = await db.execute(
        select(MedicalExam)
        .where(MedicalExam.user_id == learner.id)
        .order_by(MedicalExam.exam_date.desc())
    )
    return result.scalars().all()
